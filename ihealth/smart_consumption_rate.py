import os
import xlrd
import xlsxwriter
import openpyxl
import re


working_month='201708'
os.chdir("E:/sync360/documents/VVV项目方案VVV/智慧医疗诊间结算/统计数据/{}".format(working_month))
workbook = openpyxl.load_workbook("smart_consumption_rate.xlsx") if os.path.exists("smart_consumption_rate.xlsx") else openpyxl.workbook.Workbook()
# worksheet = workbook.get_sheet_by_name("结算率")   
try:
     worksheet = workbook.get_sheet_by_name("结算率")
except KeyError as e:
     worksheet = workbook.create_sheet("结算率")

position_list = [1,2,3,4,6]
fn_list = []
for fn in os.listdir("."):         # 逐行读取目录文件到file
    if re.fullmatch(r'.*(上报表|院).*\.(xls|xlsx)', fn) != None:
        fn_list.append(fn)
print(fn_list)
for fn in fn_list:
    book = xlrd.open_workbook(fn)
    sheet1 = book.sheet_by_index(0)
    print(sheet1.name)
    for ri in range(3, 10):
        print('row: {}'.format(ri))
        if sheet1.cell(ri, 1).ctype != 0: # 判断非空
            worksheet.cell(row=ri+1, column=1, value=sheet1.cell(ri,0).value)
            for col_idx in position_list:  # Iterate through columns
                cell_obj = sheet1.cell(ri, col_idx)  # Get cell object by row, col
                print ('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
                if cell_obj.ctype != 0 and int(cell_obj.value) >= 1:
                    worksheet.cell(row=ri+1, column=col_idx+1, value=cell_obj.value)
                else: worksheet.cell(row=ri+1, column=col_idx+1, value=0)

workbook.save("smart_consumption_rate.xlsx")

# Print all values, iterating through rows and columns
#
''' 
num_cols = sheet1.ncols   # Number of columns
for row_idx in range(0, sheet1.nrows):    # Iterate through rows
    print ('-'*40)
    print ('Row: %s' % row_idx)   # Print row number
    for col_idx in range(0, num_cols):  # Iterate through columns
        cell_obj = sheet1.cell(row_idx, col_idx)  # Get cell object by row, col
        print ('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
        print('ctype {}---{}'.format(type(cell_obj.ctype), cell_obj.ctype))
        print('value {}---{}'.format(type(cell_obj.value), cell_obj.value))
 '''