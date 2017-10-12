import os
import xlrd
import xlsxwriter
import openpyxl
import re
import sqlalchemy
import pandas as pd
import numpy as np

os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
community_request_sql = """SELECT
	x.yylsh,
	x.yymc,
	y.nnn
FROM
	XDRPARAMETER x,
	(
		SELECT DISTINCT
			(t1.transactionrequest),
			COUNT (*) AS nnn
		FROM
			(
				SELECT DISTINCT
					(T .accessionnumber) AS jch,
					T .transactionrequest
				FROM
					transactiontrack T
				WHERE
					T .transactionrequesttime >= '20170801'
				AND T .transactionrequesttime < '20170901'
			) t1
		GROUP BY
			t1.transactionrequest
	) y
WHERE
	x.yymc = y.transactionrequest
ORDER BY
	x.yylsh
"""
community_total_sql = """SELECT ADMISSIONSOURCEHOSPITAL as yymc, count(*) as mmm from EISSTUDIES WHERE STUDIESDATE >= '20170801' and STUDIESDATE < '20170901' GROUP BY ADMISSIONSOURCEHOSPITAL
"""
hospital_pair_sql = """SELECT
	x.YYLSH AS responselsh,
	y.TRANSACTIONRESPONSE,
	X1.YYLSH AS requestlsh,
	y.TRANSACTIONREQUEST,
	nnn
FROM
	(
		XDRPARAMETER x
		JOIN (
			SELECT
				t1.transactionrequest,
				T1.TRANSACTIONRESPONSE,
				COUNT (*) AS nnn
			FROM
				(
					SELECT DISTINCT
						(T .accessionnumber) AS jch,
						T .transactionrequest,
						T .TRANSACTIONRESPONSE
					FROM
						transactiontrack T
					WHERE
						T .transactionrequesttime >= '20170801'
					AND T .transactionrequesttime < '20170901'
				) t1
			GROUP BY
				T1.TRANSACTIONREQUEST,
				T1.TRANSACTIONRESPONSE
		) y ON x.yymc = y.TRANSACTIONRESPONSE
	)
JOIN XDRPARAMETER x1 ON X1.yymc = y.transactionrequest
ORDER BY
	responselsh,
	requestlsh"""


working_month='201708'
os.chdir("E:/sync360/documents/VVV项目方案VVV/智慧医疗诊间结算/统计数据/{}".format(working_month))
workbook = openpyxl.load_workbook("smk.xlsx") if os.path.exists("smk.xlsx") else openpyxl.workbook.Workbook()
# worksheet = workbook.get_sheet_by_name("结算率")
try:
    worksheet = workbook.get_sheet_by_name("医院影像")
except KeyError as e:
    worksheet = workbook.create_sheet("医院影像")


enginex = sqlalchemy.create_engine(
    'oracle://jgpt:jgpt@192.168.201.49:1521/?service_name=orcl')
result_proxy = enginex.execute(hospital_pair_sql)

rows = result_proxy.fetchall()
last_hospital = ""
r = 2
c = 1
total = 0
for row in rows:
    if row['transactionresponse'] != last_hospital:
        if last_hospital != "":
            worksheet.cell(row = r, column= c, value = last_hospital)
            worksheet.cell(row = r, column= c+1, value = '小计')
            worksheet.cell(row = r, column= c+2, value = total)
            r = r + 1
        total = 0
    worksheet.cell(row = r, column= c, value = row['transactionresponse'])
    worksheet.cell(row = r, column= c+1, value = row['transactionrequest'])
    worksheet.cell(row = r, column= c+2, value = row['nnn'])
    r = r+1    
    total = total + row['nnn']
    last_hospital = row['transactionresponse']
# 补最后一次小计
worksheet.cell(row = r, column= c, value = last_hospital)
worksheet.cell(row = r, column= c+1, value = '小计')
worksheet.cell(row = r, column= c+2, value = total)

workbook.save('smk.xlsx')

df1 = pd.read_sql(sql=community_request_sql, con=enginex)
engine2 = sqlalchemy.create_engine(
    'oracle://radinfo:pacs@192.168.201.49:1521/?service_name=orcl')
df2 = pd.read_sql(sql=community_total_sql, con=engine2)
df1 = df1.drop(labels='yylsh', axis=1)
df2 = df2.dropna()
# print(df1)
# print(df2)
df3 = pd.merge(left=df1, right=df2, on='yymc', how='outer')
df3 = df3.fillna(0)
df3['ratio']=df3['nnn']/df3['mmm']
df3 = df3.astype({'nnn':np.int64, 'mmm':np.int64})
print(df3)
try:
    worksheet = workbook.get_sheet_by_name("社区影像")
except KeyError as e:
    worksheet = workbook.create_sheet("社区影像")

i=0
for row in df3.itertuples(index=False):
    worksheet.cell(row = i+2, column= 1, value = row[0])
    worksheet.cell(row = i+2, column= 2, value = row[1])
    worksheet.cell(row = i+2, column= 3, value = row[2])
    i = i+1

workbook.save('smk.xlsx')