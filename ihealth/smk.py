import os
import xlrd
import xlsxwriter
import openpyxl
import re
import sqlalchemy


working_month='201708'
os.chdir("E:/sync360/documents/VVV项目方案VVV/智慧医疗诊间结算/统计数据/{}".format(working_month))
workbook = openpyxl.load_workbook("smk.xlsx") if os.path.exists("smk.xlsx") else openpyxl.workbook.Workbook()
# worksheet = workbook.get_sheet_by_name("结算率")   
try:
     worksheet = workbook.get_sheet_by_name("市民卡")
except KeyError as e:
     worksheet = workbook.create_sheet("市民卡")

os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
enginex = sqlalchemy.create_engine(
    'oracle://xiaohu:82626296@192.168.201.77:1521/?service_name=pdborcl')
result_proxy = enginex.execute(r"SELECT * from SMK_STATISTIC sta join ORGANIZATION1 org on ORG.UUID = STA.HOSPITAL_ID where STA.REPORTING_MONTH = TO_DATE(201708, 'yyyyMM')")

rows = result_proxy.fetchall()
for r in rows:
    for i in range(2, 18):
        print("name: {}-{}".format(r['name'], worksheet.cell(row=i, column=2).value))
        if r['name'] == worksheet.cell(row=i, column=2).value:
            worksheet.cell(row=i, column=3, value=r['healthy_card_issued'] if r['healthy_card_issued']!=None else 0)
            worksheet.cell(row=i, column=4, value=r['smart_consumption_times'] if r['smart_consumption_times']!=None else 0)
            worksheet.cell(row=i, column=5, value=r['smart_consumption_amount'] if r['smart_consumption_amount']!=None else 0)
            worksheet.cell(row=i, column=6, value=r['counter_recharge_times'] if r['counter_recharge_times']!=None else 0)
            worksheet.cell(row=i, column=7, value=r['counter_recharge_amount'] if r['counter_recharge_amount']!=None else 0)
            worksheet.cell(row=i, column=8, value=r['machine_recharge_times'] if r['machine_recharge_times']!=None else 0)
            worksheet.cell(row=i, column=9, value=r['machine_recharge_amount'] if r['machine_recharge_amount']!=None else 0)
workbook.save('smk.xlsx')

